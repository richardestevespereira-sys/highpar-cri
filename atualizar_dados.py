# -*- coding: utf-8 -*-
"""
ATUALIZADO1 — Extrator GENÉRICO v3: lê qualquer consolidado padrão de CRI (via
aba _SCHEMA) e gera/atualiza o banco de dados do portal (site/dados.json e
site/dados.js — o .js permite abrir o portal SEM servidor, direto do arquivo).

v3 (novidades):
  * Séries com Código IF e Quantidade emitida (Config, colunas C e D).
  * GARANTIAS da Config (bloco cfg_gar via _SCHEMA) + enriquecimento pelo
    cadastro/<id>.json (garantias detalhadas, COVENANTS e POSIÇÃO HIGHPAR).
  * Validação por emissão: relatório de caixinhas preenchidas/faltantes no
    console e embutido no banco (portal mostra o que falta).
  * Radar de Risco continua 100% calculado pelo portal (nunca na planilha).

Uso:
  python atualizar_dados.py                       -> varre "Consolidado - *.xlsx"
  python atualizar_dados.py arq1.xlsx arq2.xlsx   -> só os citados
"""
import sys, os, json, glob, re, datetime as dt, unicodedata
from statistics import median
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string as CI

def d2s(x): return x.strftime('%Y-%m-%d') if isinstance(x,(dt.date,dt.datetime)) else None

def num(x):
    if isinstance(x, bool): return None
    if isinstance(x,(int,float)):
        return float(x) if float('-inf') < float(x) < float('inf') else None
    if isinstance(x,str):
        t = x.strip().replace(' ', '').replace(' ', '')
        if not t: return None
        # Aceita 1.234,56 e 1,234.56 sem alterar valores decimais validos.
        if ',' in t and '.' in t:
            if t.rfind(',') > t.rfind('.'):
                t = t.replace('.', '').replace(',', '.')
            else:
                t = t.replace(',', '')
        elif ',' in t:
            t = t.replace('.', '').replace(',', '.')
        try:
            v = float(t)
            return v if float('-inf') < v < float('inf') else None
        except ValueError: return None
    return None

def slug(s):
    s = re.sub(r'[^a-zA-Z0-9]+','_', str(s).strip().lower())
    return s.strip('_')

def texto_normalizado(x):
    """Texto comparavel entre layouts, sem depender de acentos ou caixa."""
    return unicodedata.normalize('NFKD', str(x or '')).encode('ascii', 'ignore').decode().strip().lower()

def nome_faixa_atraso(faixa=None, dias=None):
    """Converte faixas heterogeneas do servicer para as chaves do portal."""
    txt = texto_normalizado(faixa)
    nums = [int(n) for n in re.findall(r'\d+', txt)]
    if 'acima' in txt or 'maior' in txt or ('+' in txt and nums):
        return 'Inad Acima de 360 dias'
    if len(nums) >= 2:
        inicio, fim = nums[0], nums[1]
    else:
        atraso = num(dias)
        if atraso is None or atraso <= 0:
            return None
        inicio, fim = ((1, 30) if atraso <= 30 else (31, 60) if atraso <= 60 else
                       (61, 90) if atraso <= 90 else (91, 120) if atraso <= 120 else
                       (121, 180) if atraso <= 180 else (181, 360) if atraso <= 360 else (361, 99999))
    if fim <= 30: return 'Inad Ate 30 dias'
    if inicio <= 60 and fim <= 60: return 'Inad De 31 a 60 dias'
    if inicio <= 90 and fim <= 90: return 'Inad De 61 a 90 dias'
    if inicio <= 120 and fim <= 120: return 'Inad De 91 a 120 dias'
    if inicio <= 180 and fim <= 180: return 'Inad De 121 a 180 dias'
    if inicio <= 360 and fim <= 360: return 'Inad De 181 a 360 dias'
    return 'Inad Acima de 360 dias'

def complementar_kpi_inadimplencia(wb, kpi):
    """Resume BD_Inadimplencia quando o aging ainda nao veio no BD_KPI."""
    if 'BD_Inadimplencia' not in wb.sheetnames:
        return
    ws = wb['BD_Inadimplencia']
    headers = [texto_normalizado(c.value) for c in ws[1]]

    def coluna(*padroes):
        return next((i for i, h in enumerate(headers) if any(re.search(p, h) for p in padroes)), None)

    c_data = coluna(r'^data do relatorio$', r'^safra$')
    c_faixa = coluna(r'periodo de atraso', r'faixa.*atraso', r'^faixa$')
    c_dias = coluna(r'dias de atraso')
    c_valor = coluna(r'^valor em atraso$', r'valor.*inadimpl')
    c_parcelas = coluna(r'parcelas em aberto')
    if c_data is None or c_valor is None:
        return
    totais, parcelas = {}, {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = row[c_data] if c_data < len(row) else None
        valor = num(row[c_valor]) if c_valor < len(row) else None
        if not isinstance(data, (dt.date, dt.datetime)) or valor is None or valor <= 0:
            continue
        mes = d2s(data)[:7]
        faixa = nome_faixa_atraso(row[c_faixa] if c_faixa is not None and c_faixa < len(row) else None,
                                  row[c_dias] if c_dias is not None and c_dias < len(row) else None)
        if not faixa:
            continue
        faixas = totais.setdefault(mes, {})
        faixas[faixa] = faixas.get(faixa, 0) + valor
        if c_parcelas is not None and c_parcelas < len(row):
            parcelas[mes] = parcelas.get(mes, 0) + (num(row[c_parcelas]) or 0)
    for mes, faixas in totais.items():
        destino = kpi.setdefault(mes, {})
        for faixa, valor in faixas.items():
            destino.setdefault(faixa, round(valor, 2))
        destino.setdefault('Inadimplencia (R$)', round(sum(faixas.values()), 2))
        if parcelas.get(mes):
            destino.setdefault('Inadimplencia (parcelas)', round(parcelas[mes], 0))

def _cabecalhos(ws):
    return [texto_normalizado(c.value) for c in ws[1]]

def _coluna(headers, *padroes):
    return next((i for i, h in enumerate(headers) if any(re.search(p, h) for p in padroes)), None)

def _mes_data(x):
    return d2s(x)[:7] if isinstance(x, (dt.date, dt.datetime)) else None

def extrair_analiticos_inadimplencia(wb, kpi):
    """Gera series reutilizaveis sem misturar mes de caixa, vencimento e snapshot."""
    rastro, faixas = [], []

    # Rastro por coorte de vencimento no ultimo snapshot comum. Exige uma
    # classificacao explicita do saldo aberto para nao presumir inadimplencia.
    if 'BD_Recebimento' in wb.sheetnames and 'BD_Receber' in wb.sheetnames:
        wr, wa = wb['BD_Recebimento'], wb['BD_Receber']
        hr, ha = _cabecalhos(wr), _cabecalhos(wa)
        rr = _coluna(hr, r'^data do relatorio$', r'^safra')
        rv = _coluna(hr, r'^data vencimento$', r'^dt\.? venc', r'^vencimento$')
        rval = _coluna(hr, r'^valor corrigido$', r'^valor parcela$', r'^valor pago$')
        ar = _coluna(ha, r'^data do relatorio$', r'^safra')
        av = _coluna(ha, r'^data vencimento$', r'^dt\.? venc', r'^vencimento$')
        aval = _coluna(ha, r'^valor corrigido$', r'^valor parcela$', r'^valor em atraso$', r'^valor$')
        acl = _coluna(ha, r'^classificacao$', r'^status$', r'^situacao$')
        if None not in (rr, rv, rval, ar, av, aval, acl):
            sr = {_mes_data(row[rr]) for row in wr.iter_rows(min_row=2, values_only=True)}
            sa = {_mes_data(row[ar]) for row in wa.iter_rows(min_row=2, values_only=True)}
            comuns = sorted((sr & sa) - {None})
            if comuns:
                ref = comuns[-1]
                recebido, vencido = {}, {}
                for row in wr.iter_rows(min_row=2, values_only=True):
                    if _mes_data(row[rr]) != ref: continue
                    mes, valor = _mes_data(row[rv]), num(row[rval])
                    if mes and mes < ref and valor is not None and valor > 0:
                        recebido[mes] = recebido.get(mes, 0) + valor
                for row in wa.iter_rows(min_row=2, values_only=True):
                    if _mes_data(row[ar]) != ref: continue
                    status = texto_normalizado(row[acl])
                    if not re.search(r'vencid|atras|inadimpl', status): continue
                    mes, valor = _mes_data(row[av]), num(row[aval])
                    if mes and mes < ref and valor is not None and valor > 0:
                        vencido[mes] = vencido.get(mes, 0) + valor
                for mes in sorted(set(recebido) | set(vencido)):
                    rec, ina = recebido.get(mes, 0), vencido.get(mes, 0)
                    total = rec + ina
                    if total > 0:
                        rastro.append({'m': mes, 'recebido': round(rec / total, 6),
                                       'inadimplente': round(ina / total, 6), 'as_of': ref})

    # Curvas cumulativas >30/>60/>90 por snapshot, sobre o saldo da carteira.
    if 'BD_Inadimplencia' in wb.sheetnames:
        ws, h = wb['BD_Inadimplencia'], _cabecalhos(wb['BD_Inadimplencia'])
        cd = _coluna(h, r'^data do relatorio$', r'^safra')
        cdi = _coluna(h, r'^dias de atraso$', r'^dias mais atrasad')
        cv = _coluna(h, r'^valor em atraso$', r'valor.*inadimpl')
        cs = _coluna(h, r'^saldo devedor$')
        cp = _coluna(h, r'participacao.*saldo devedor')
        if None not in (cd, cdi, cv):
            bruto, implicitos = {}, {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                mes, dias, valor = _mes_data(row[cd]), num(row[cdi]), num(row[cv])
                if not mes or dias is None or valor is None or valor <= 0: continue
                x = bruto.setdefault(mes, {'p30': 0, 'p60': 0, 'p90': 0})
                if dias > 30: x['p30'] += valor
                if dias > 60: x['p60'] += valor
                if dias > 90: x['p90'] += valor
                if cs is not None and cp is not None:
                    saldo, part = num(row[cs]), num(row[cp])
                    if saldo and part and part > 0:
                        implicitos.setdefault(mes, []).append(saldo / part)
            for mes in sorted(bruto):
                base = next((num(v) for nome, v in (kpi.get(mes) or {}).items()
                             if re.search(r'saldo devedor.*carteira', texto_normalizado(nome)) and num(v)), None)
                if not base and implicitos.get(mes): base = median(implicitos[mes])
                if not base or base <= 0: continue
                faixas.append({'m': mes, 'p30': round(bruto[mes]['p30'] / base, 6),
                               'p60': round(bruto[mes]['p60'] / base, 6),
                               'p90': round(bruto[mes]['p90'] / base, 6),
                               'base': round(base, 2)})
    return rastro, faixas

def mensalizar_recebimentos(receb):
    """Consolida movimentos validos por mes para o portal."""
    mensal = {}
    for r in receb:
        d, v = r.get('d'), num(r.get('v'))
        if not (isinstance(d, str) and re.fullmatch(r'\d{4}-\d{2}-\d{2}', d)):
            continue
        if v is None or v <= 0:
            continue
        mensal[d[:7]] = mensal.get(d[:7], 0) + v
    return [{'m': m, 'v': round(mensal[m], 2)} for m in sorted(mensal)]

def extrair(path):
    wb = load_workbook(path, read_only=False, data_only=False)
    if '_SCHEMA' not in wb.sheetnames:
        print(f'  AVISO: {os.path.basename(path)} sem _SCHEMA — ignorado.')
        return None
    sc = json.loads(''.join(c[0].value for c in wb['_SCHEMA'].iter_rows(min_row=2, max_col=1) if c[0].value))
    cfg = wb['Config']
    nome   = cfg['B3'].value or 'Emissão'
    titulo = cfg['B4'].value or 'CRI'

    # características (rótulo A, valor B, linhas 3..14)
    carac = []
    for r in range(3, 15):
        k, v = cfg.cell(row=r, column=1).value, cfg.cell(row=r, column=2).value
        if k is not None and v is not None:
            carac.append({'k': str(k), 'v': d2s(v) or (num(v) if num(v) is not None else str(v))})

    # séries (com código IF, quantidade emitida e taxa)
    series = []
    srow = sc.get('cfg_series_row', 18)
    for i in range(int(sc.get('n_series', 1))):
        r = srow + i
        cod_if = cfg.cell(row=r, column=3).value
        series.append({'n': cfg.cell(row=r, column=1).value,
                       'nome': cfg.cell(row=r, column=2).value,
                       'cod_if': str(cod_if) if cod_if not in (None, '') else None,
                       'qtd': num(cfg.cell(row=r, column=4).value),
                       'taxa': num(cfg.cell(row=r, column=7).value)})

    # players da operação (Config, bloco OPERAÇÃO via _SCHEMA)
    players = []
    prow, pn = sc.get('cfg_ops_row'), sc.get('cfg_ops_n', 0)
    if prow:
        for i in range(int(pn)):
            r = prow + i
            papel, quem = cfg.cell(row=r, column=1).value, cfg.cell(row=r, column=2).value
            if papel and quem: players.append({'papel': str(papel), 'nome': str(quem)})

    # garantias declaradas na Config (bloco GARANTIAS via _SCHEMA)
    garantias_cfg = []
    grow, gn = sc.get('cfg_gar_row'), sc.get('cfg_gar_n', 0)
    if grow:
        for i in range(int(gn)):
            nome_g = cfg.cell(row=grow + i, column=1).value
            if nome_g:
                garantias_cfg.append({'nome': str(nome_g), 'tipo': None,
                                      'descricao': None, 'status': 'Vigente',
                                      'fonte': 'Config'})

    # KPIs por safra (todas as métricas do BD_KPI, sem hardcode)
    kpi = {}
    if 'BD_KPI' in wb.sheetnames:
        for row in wb['BD_KPI'].iter_rows(min_row=2, max_col=3):
            s, m, v = [c.value for c in row]
            v = num(v)
            if s and m and v is not None:
                kpi.setdefault(d2s(s)[:7], {})[str(m)] = round(v, 4)
    complementar_kpi_inadimplencia(wb, kpi)
    inad_trail, inad_faixas = extrair_analiticos_inadimplencia(wb, kpi)

    # extrato: recebimentos da carteira e pagamentos ao CRI (flags do _SCHEMA)
    receb, pag_cri = [], {}
    flags = sc.get('c', {}).get('flags', {})
    fcart = next((k for k in flags if 'carteira' in k.lower()), None)
    fpag  = next((k for k in flags if 'pagamento cri' in k.lower()), None)
    if 'C' in wb.sheetnames and (fcart or fpag):
        cols = [CI(flags[f]) for f in (fcart, fpag) if f]
        for row in wb['C'].iter_rows(min_row=sc['c']['first_row'], max_col=max(cols+[9])):
            v = [c.value for c in row]
            if not isinstance(v[0],(dt.date,dt.datetime)): continue
            val = num(v[4]) if len(v) > 4 else None
            if val is None: continue
            if fcart and v[CI(flags[fcart])-1] == 1 and val > 0:
                receb.append({'d': d2s(v[0]), 'v': round(val, 2)})
            if fpag and v[CI(flags[fpag])-1] == 1:
                m = d2s(v[0])[:7]
                pag_cri[m] = round(pag_cri.get(m, 0) + abs(val), 2)

    # integralizações (aportes do investidor): data B, qtd D, série F, PU G
    integ = []
    it = sc.get('integ', {})
    if 'Integralização' in wb.sheetnames and it:
        ws = wb['Integralização']
        for r in range(it.get('first_row', 3), it.get('first_row', 3) + it.get('rows', 300)):
            d = ws.cell(row=r, column=2).value
            qtd, serie, pu = num(ws.cell(row=r, column=4).value), ws.cell(row=r, column=6).value, num(ws.cell(row=r, column=7).value)
            if isinstance(d,(dt.date,dt.datetime)) and qtd and pu:
                integ.append({'d': d2s(d), 'serie': serie, 'v': round(qtd*pu, 2)})

    # carteira ESPERADA por mês (BD_Receber: acha colunas de vencimento e valor)
    receber_prev = {}
    if 'BD_Receber' in wb.sheetnames:
        ws = wb['BD_Receber']
        hdr = [str(c.value or '') for c in ws[1]]
        crel = next((i for i,h in enumerate(hdr) if re.search(r'relat', h, re.I)), None)
        cd   = next((i for i,h in enumerate(hdr) if re.search(r'venc', h, re.I)), None)
        cv   = next((i for i,h in enumerate(hdr) if re.search(r'valor', h, re.I)), None)
        if cd is not None and cv is not None:
            for r in ws.iter_rows(min_row=2, values_only=True):
                d, v = r[cd], num(r[cv])
                if not (isinstance(d,(dt.date,dt.datetime)) and v): continue
                # BD_Receber acumula snapshots mensais: usa só o snapshot do
                # próprio mês (Data do Relatório no mesmo mês do vencimento)
                if crel is not None:
                    rel = r[crel]
                    if not (isinstance(rel,(dt.date,dt.datetime)) and (rel.year,rel.month)==(d.year,d.month)):
                        continue
                m = d2s(d)[:7]
                receber_prev[m] = round(receber_prev.get(m, 0) + v, 2)

    # obra
    obra = []
    ob = sc.get('obra')
    if ob and 'Obra' in wb.sheetnames:
        ws = wb['Obra']; cols = ob.get('cols', {})
        cm, cp, cmed = CI(cols.get('mes','A')), CI(cols.get('pct','B')), CI(cols.get('medicao','E'))
        for r in range(ob['first_row'], ob['first_row'] + ob['rows']):
            m = ws.cell(row=r, column=cm).value
            pct, med = num(ws.cell(row=r, column=cp).value), num(ws.cell(row=r, column=cmed).value)
            if m and (pct is not None or med is not None):
                obra.append({'m': d2s(m)[:7], 'pct': round(pct or 0, 4), 'med': round(med or 0, 2)})
    orcamento = num(cfg['B9'].value) or 0

    # limiares do farol (Config) — só os LIMITES; o cálculo é do portal
    farol_regras = []
    frow, fn = sc.get('cfg_farol_row'), sc.get('cfg_farol_n', 0)
    if frow:
        for i in range(int(fn)):
            r = frow + i
            ind = cfg.cell(row=r, column=1).value
            if not ind: continue
            farol_regras.append({'indicador': str(ind),
                                 'dir': str(cfg.cell(row=r, column=3).value or '>='),
                                 'verde': num(cfg.cell(row=r, column=4).value),
                                 'amarelo': num(cfg.cell(row=r, column=5).value)})

    receb_mensal = mensalizar_recebimentos(receb)
    return {'id': slug(f"{titulo}_{nome}"), 'nome': str(nome), 'titulo': str(titulo),
            'arquivo': os.path.basename(path), 'gerado': dt.date.today().isoformat(),
            'carac': carac, 'series': series, 'players': players, 'kpi': kpi,
            'receb': receb, 'receb_mensal': receb_mensal, 'pag_cri': pag_cri, 'integ': integ,
            'receber_prev': receber_prev, 'obra': obra, 'orcamento': orcamento,
            'inad_trail': inad_trail, 'inad_faixas': inad_faixas,
            'farol_regras': farol_regras, 'garantias': garantias_cfg}

def aplicar_cadastro(e):
    """Enriquece a emissão com cadastro/<id>.json: garantias detalhadas,
    covenants (obrigações verificáveis) e posição Highpar. O cadastro NUNCA
    sobrescreve dados extraídos da planilha — só complementa/detalha."""
    caminho = os.path.join('cadastro', f"{e['id']}.json")
    if not os.path.exists(caminho):
        e.setdefault('covenants', [])
        e.setdefault('posicao_highpar', None)
        e.setdefault('status', 'Ativa')
        return e
    try:
        with open(caminho, encoding='utf-8') as fh:
            cad = json.load(fh)
    except (OSError, json.JSONDecodeError) as exc:
        print(f"  AVISO: cadastro invalido ({caminho}): {exc}")
        return e
    if cad.get('garantias'):
        e['garantias'] = cad['garantias']   # versão detalhada vence a lista simples da Config
    e['covenants'] = cad.get('covenants', [])
    e['posicao_highpar'] = cad.get('posicao_highpar')
    e['status'] = cad.get('status', 'Ativa')
    e['codigo_emissao'] = cad.get('codigo_emissao')
    e['dossie'] = cad.get('dossie')
    return e

CAIXINHAS = [
    ('carac', 'Características'), ('series', 'Séries (IF/qtd/taxa)'),
    ('players', 'Players da operação'), ('kpi', 'KPIs por safra'),
    ('receb_mensal', 'Recebimentos mensais'), ('pag_cri', 'Pagamentos ao CRI'),
    ('integ', 'Integralizações'), ('receber_prev', 'Carteira esperada'),
    ('obra', 'Obra'), ('inad_faixas', 'Inadimplência (faixas)'),
    ('farol_regras', 'Limiares do radar'), ('garantias', 'Garantias'),
    ('covenants', 'Covenants'),
]

def validar_caixinhas(e):
    """Marca o que está preenchido/faltando (impresso e embutido no banco)."""
    v = {}
    for chave, rotulo in CAIXINHAS:
        val = e.get(chave)
        ok = bool(val) and (len(val) > 0 if hasattr(val, '__len__') else True)
        if chave == 'series' and ok:
            ok = all(s.get('nome') for s in val)
            sem_if = [s['nome'] for s in val if not s.get('cod_if')]
            sem_qtd = [s['nome'] for s in val if not s.get('qtd')]
            if sem_if or sem_qtd:
                v['series_obs'] = ('sem IF: ' + ', '.join(map(str, sem_if)) if sem_if else '') + \
                                  ((' | ' if sem_if and sem_qtd else '') +
                                   ('sem qtd: ' + ', '.join(map(str, sem_qtd)) if sem_qtd else ''))
        v[chave] = ok
    pos = e.get('posicao_highpar') or {}
    v['posicao_highpar'] = any(s.get('qtd_detida') for s in pos.get('por_serie', []))
    e['validacao'] = v
    faltas = [rot for ch, rot in CAIXINHAS if not v.get(ch)]
    if not v['posicao_highpar']:
        faltas.append('Posição Highpar (preencher no cadastro)')
    print('  caixinhas ok:', sum(1 for ch, _ in CAIXINHAS if v.get(ch)), '/', len(CAIXINHAS),
          ('| faltando: ' + '; '.join(faltas) if faltas else '| completo'))
    return e

def ler_base_existente(path):
    if not os.path.exists(path):
        return {'emissoes': []}
    try:
        with open(path, encoding='utf-8') as fh:
            base = json.load(fh)
    except (OSError, json.JSONDecodeError) as exc:
        raise RuntimeError(f'Nao foi possivel ler {path}: {exc}') from exc
    if not isinstance(base, dict) or not isinstance(base.get('emissoes', []), list):
        raise RuntimeError(f'{path} nao possui o formato esperado de banco do portal.')
    return base

def arquivos_padrao():
    ignorar = ('backup', 'teste', 'template')
    return [f for f in glob.glob('Consolidado - *.xlsx')
            if not any(tag in os.path.basename(f).lower() for tag in ignorar)]

def validar_emissao(e, path):
    if not e['id']:
        raise ValueError(f'{path}: Config!B3 e Config!B4 devem identificar a emissao.')
    if not e['kpi'] and not e['receb_mensal'] and not e['integ'] and not e['obra']:
        raise ValueError(f'{path}: consolidado sem safras ou movimentos validos.')

def main():
    args = sys.argv[1:]
    arquivos = args or arquivos_padrao()
    if not arquivos:
        raise SystemExit('Nenhum arquivo "Consolidado - *.xlsx" valido foi encontrado.')

    destino = os.environ.get('DADOS_JSON') or (os.path.join('site', 'dados.json') if os.path.isdir('site') else 'dados.json')
    base = ler_base_existente(destino)
    por_id = {e.get('id'): e for e in base['emissoes'] if e.get('id')}
    atualizadas, novas = [], []
    for f in sorted(arquivos):
        print('Lendo:', f)
        e = extrair(f)
        if e:
            validar_emissao(e, f)
            e = aplicar_cadastro(e)
            e = validar_caixinhas(e)
            if e['id'] in por_id:
                atualizadas.append(e['id'])
                acao = 'atualizada'
            else:
                novas.append(e['id'])
                acao = 'nova'
            por_id[e['id']] = e
            print(f"  ok: {e['titulo']} {e['nome']} | safras {len(e['kpi'])} | receb {len(e['receb'])} | integ {len(e['integ'])} | prev {len(e['receber_prev'])} | players {len(e['players'])}")
            print(f"  emissao {acao}: {e['id']} | meses recebidos {len(e['receb_mensal'])}")
    os.makedirs('site', exist_ok=True)
    out = {'gerado': dt.datetime.now().isoformat(timespec='minutes'),
           'emissoes': [por_id[k] for k in sorted(por_id)]}
    with open(destino, 'w', encoding='utf-8') as fh:
        json.dump(out, fh, ensure_ascii=False)
    # dados.js: mesmo banco como script — o portal abre direto do disco (file://)
    destino_js = os.path.splitext(destino)[0] + '.js'
    with open(destino_js, 'w', encoding='utf-8') as fh:
        fh.write('window.DADOS = ')
        json.dump(out, fh, ensure_ascii=False)
        fh.write(';\n')
    print('Banco atualizado:', destino, 'e', destino_js, '|', len(out['emissoes']), 'emissões',
          f'| atualizadas {len(atualizadas)} | novas {len(novas)}')

if __name__ == '__main__':
    main()
